#!/usr/bin/env python
# coding: utf-8

# In[ ]:


####### DESCRIPTION#########
#This code was written for uploading data and modified image files in a customised database, the overall idea behind writing
#this code was to reduce manual intervention by using the power and rich libraries of python in order to increase the data
#submission speed with increased efficiency.

#Input: Excel sheet containing data and path of images present in local system/server.
#Output: Returning the excel sheet populated with the above data along with submission entry id from the db and upload status.


### Python libraries used for achieving objective#########

import xlrd as xl
from xlutils.copy import copy
import xlwt as wt
import regex as re
import operator
import time
from datetime import datetime
import os
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
import smtplib
import openpyxl
from openpyxl import load_workbook
import xml.etree.ElementTree as ET

## Module for modifying image files.

def texonImage(imgfilepath, PPB_ID, PPB_NAME):
    #    Image format extension
    imgformat = ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif', '.rgb', '.pbm', '.pgm', '.ppm', '.rast', '.xbm', '.webp', '.exr']
    
    #    Validating the image file
    if imgfilepath == "" or imgfilepath == None or os.path.exists(imgfilepath) == False or imgfilepath.lower().endswith((tuple(imgformat))) == False:
        print("Invalid")    
    else:      
    #   Parsing the image from input file
        image = Image.open(imgfilepath).convert('RGB')

    #    initialise the drawing context with
    #    the image object as background
        draw = ImageDraw.Draw(image)

        #    create font object with the font file and specify
        #    desired size
        font = ImageFont.truetype('arial.ttf', size=15)
    #    starting position of the message
        (x, y) = (45, 2)    
    #    Validating PPB_NAME exits
        if PPB_NAME == None:
            message = PPB_ID
        else:
            message = PPB_ID + " | " + PPB_NAME    
        color = 'rgb(0, 0, 0)' # black color
    #    draw the message on the background
        draw.text((x, y), message, fill=color, font=font)    
    #    Saving the image
        image.save(imgfilepath)
############################################################################################################################

##Uploading Image module in demodb.
def upload_image(path,ID,Name):
    
    baseURL="https://www.demodb.com"
    if ID == "":                         #Uploading Image file using Product name incase Product id is not present.
        img_ref=str(Name)+str(path)
        image_URL=str(baseURL) + "/dbImageUpload/upload?folder="+str(Name)+"filename="+str(path)+"secret=YYYYYYYYYY"
    else:                               #Uploading Image file using Product ID whether product name is present or not.
        img_ref=str(ID)+str(path)
        image_URL=str(baseURL)+"/dbImageUpload/upload?folder="+str(ID)+"filename="+str(path)+"secret=YYYYYYYYYY"
        
    print (image_URL)
    
###########################################################################################################################

## Parsing XML response got after pushing data in demodb and fetching the submission id,
## if the push data is completed without any errors.

def parse_xml(path):
    
    with os.scandir(path) as entries:
        for file in entries:
            if file.is_file():
                name=file.name
                if name[-4:] == ".xml":
                    file_name=path+"//"+name
                    print (file_name)
                
                    tree = ET.parse(file_name)
                    root = tree.getroot()
                    pattern=re.compile(r"PR-\d+")
                    #print (pattern)
                    cnt=0
                    for child in root:
                        for results in child:
                            cnt=cnt+1
                            if cnt == 1:
                                operation=[]
                                for v in results.attrib.values():
                                    operation.append(v)
                            else:
                                cnt=0
                            for entity in results:
                                if pattern.search(entity.text):
                                    operation.append(entity.text)

                        os.remove(path)
                    print (operation)
###########################################################################################################################



SHEET_PATH=("C://Users//Syatems//Desktop//db//Program//TEST_SHEET2.xlsx")
tmp_path="C://Users//Syatems//Desktop//db//Program//TMP//"

wb=xl.open_workbook(SHEET_PATH)

pxl=load_workbook(SHEET_PATH)

ws=pxl["Sheet1"]

#sheet_obj=pxl.active

sheet=wb.sheet_by_index(0)

nrows=sheet.nrows
ncols=sheet.ncols
col_list={}
ind=[]
#print (nrows,ncols,sep="\t")

if nrows >3:
    MT=sheet.cell_value(0,2)
    head=[]
    for colname in range(ncols):
       columns=re.sub('\[.*\]','',sheet.cell_value(2,colname))
       col_list[columns]=colname
       head.append(columns)
    head="\t".join(head)
    #print (head)
    
    #col_list.append(sheet.cell_value(2,colname))
    
    idx_id=col_list['PPB-ID']        #getting index location for PPB-ID
    idx_name=col_list['PPB Name']    #getting index location for PPB Name
    idx_il=col_list['Image Location']       #getting index location for Image Location
    
    if 'PPB-ID' in col_list and 'PPB Name' in col_list and 'Image Location' in col_list:
               
        for r in range(3,nrows):
            pattern=re.compile(r'(PPB-\d+)')
            ID=sheet.cell_value(r,col_list.get('PPB-ID'))
            Name=sheet.cell_value(r,col_list.get('PPB Name'))
            if pattern.search(ID) or (ID == "" and Name != ""):
                
                img_loc=sheet.cell_value(r,col_list.get('Image Location'))
                #print (ID,Name,img_loc,sep="\t")
                test=os.path.isfile(img_loc)
                #test=bool(test)
                print ("TEST RESULT IS", test)
                
                
                if img_loc != "" and  test == True:
                    print ("Image file is there")
                    now=datetime.now()
                    timestamp=datetime.timestamp(now)
                    texonImage(img_loc, ID, Name)
                    upload_image(img_loc,ID,Name)
                    h2=""
                    h1="LR ID"+ "\t"+"LR Name"+ "\t"+ "Analyzed Entity Type" + "\t"+ "Analyzed Entity ID"+ "\t" + "Analyzed Entity Name"+ "\t"+ "Laboratory Result Type" + "\t"+ "Measurement Type"
                    for k,v in col_list.items():
                        if v > 1 and v < idx_id: 
                            #print (k,v,sep="\t")
                            h2=h2+"\t"+k
                    tsv_header=h1+h2+"\t"+"Reference to Image-Original"
                    #print (tsv_header)
                    TSVfilePath=str(tmp_path)+ "dbImageUploaderTMP"+ str(timestamp) + ".tsv"
                    TSV= open(TSVfilePath,"w")
                    TSV.write(tsv_header+"\n")
                    LR_ID=["" if sheet.cell_value(r,col_list.get('LR-ID')) == "" else sheet.cell_value(r,col_list.get('LR-ID'))]
                    LR_ID="".join(LR_ID)
                    #print (LR_ID)
                    data1=LR_ID+"\t\t"+"Protein Purification Batch"+"\t"+str(ID)+"\t"+Name+"\t\t"+MT+"\t"
                    
                    data2=[]
                    for c in range(2,idx_id):
                        data2.append(sheet.cell_value(r,c))
                    data2="\t".join(map(str,data2))
                    
                    tsv_data=data1+data2+"\t"+img_loc
                    TSV.write(tsv_data)                    
                    print (tsv_data)
                    TSV.close()
                    #parse_xml(tmp_path)
                
                    ws.cell(row=(r+1),column=(idx_il+2),value="VALID")
                    ws.cell(row=(r+1),column=(idx_il+3),value="DONE")
                    ws.cell(row=(r+1),column=(idx_il+4),value="DONE")
                    
                elif img_loc == "":
                    print ("Image file is there")
                    now=datetime.now()
                    timestamp=datetime.timestamp(now)
                    upload_image(img_loc,ID,Name)
                    h2=""
                    h1="LR ID"+ "\t"+"LR Name"+"\t"+"Analyzed Entity Type" + "\t"+ "Analyzed Entity ID"+ "\t" + "Analyzed Entity Name"+ "\t"+ "Laboratory Result Type" + "\t"+ "Measurement Type"
                    for k,v in col_list.items():
                        if v > 1 and v < idx_id: 
                            #print (k,v,sep="\t")
                            h2=h2+"\t"+k
                    tsv_header=h1+h2+"\t"+"Reference to Image-Original"
                    #print (tsv_header)
                    TSVfilePath=str(tmp_path)+ "dbImageUploaderTMP"+ str(timestamp) + ".tsv"
                    TSV= open(TSVfilePath,"w")
                    TSV.write(tsv_header+"\n")
                    LR_ID=["" if sheet.cell_value(r,col_list.get('LR-ID')) == "" else sheet.cell_value(r,col_list.get('LR-ID'))]
                    LR_ID="".join(LR_ID)
                    #print (LR_ID)
                    data1=str(LR_ID)+"\t\t"+"Protein Purification Batch"+"\t"+str(ID)+"\t"+str(Name)+"\t\t"+str(MT)+"\t"
                    
                    data2=[]
                    for c in range(2,idx_id):
                        data2.append(sheet.cell_value(r,c))  
                    data2="\t".join(map(str,data2))
                    
                    tsv_data=data1+data2+"\t"+img_loc
                    TSV.write(tsv_data)                    
                    print (tsv_data)
                    TSV.close()
                    #parse_xml(tmp_path)
                    ws.cell(row=(r+1),column=(idx_il+2),value="VALID")
                    ws.cell(row=(r+1),column=(idx_il+3),value="DONE")
                    ws.cell(row=(r+1),column=(idx_il+4),value="DONE")
                                  
                else:
                    print ("INVALID IMAGE")
                    ws.cell(row=(r+1),column=(idx_il+2),value="ERROR")
                    ws.cell(row=(r+1),column=(idx_il+3),value="")
                    ws.cell(row=(r+1),column=(idx_il+4),value="")
                    
            
            elif ID == "" and Name == "":
                
                ws.cell(row=(r+1),column=(idx_il+2),value="")
                ws.cell(row=(r+1),column=(idx_il+3),value="")
                ws.cell(row=(r+1),column=(idx_il+4),value="PPB-ID and/or PPB Name must be specified!")
                
            else:
                #print ("OK3")
                ws.cell(row=(r+1),column=(idx_il+2),value="ERROR")
                ws.cell(row=(r+1),column=(idx_il+3),value="")
                ws.cell(row=(r+1),column=(idx_il+4),value="")
                
                #print ("OK")
        pxl.save(path)
    else:
        print("Template error")
else:
    print ("Data error")

