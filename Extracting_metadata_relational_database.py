#!/usr/bin/env python
# coding: utf-8

# In[2]:


#Description

#Input file:A text file having barcode id of labware.
#Outfile file:An excel sheet having all the required metadata from Oracle and Relational database related to that barcode
#along with the rack view.

"""
The following code extarcts all he required metadata related to a particular sample id from Oracle and Customised database
and returns an excel sheet which will containg all the metadata along with rack view of data which will help in understanding
sample and its metadata arrangement.
"""

#Modules used for this objective.

import warnings
warnings.filterwarnings('ignore')
import cx_Oracle
import pandas as pd
import sys
import openpyxl as op
from openpyxl import Workbook
import requests
from requests.auth import HTTPBasicAuth
import os
import io
import numpy as np
import json
import time
from datetime import datetime

#Module for connecting to oracle database and fetching required metedata from table view related to specific sample id.

def oracleconn(data):
    try:
        ids={'labware':data['samid']} #storing sample id which to be useed while querying database.
        connstr = 'username/password@hostname:port/servicename'  # creating connection string for oracle database
        conn = cx_Oracle.connect(connstr)
        cur=conn.cursor()
        
        try:
            query=cur.execute('select * from view_table_name Where sample_barcode=:labware',ids) #querying database for metadata.
            db_info=query.fetchone()
            db_info=list(db_info)
        except:
            print ("{} not found in database,hence passing".format(ids))
            pass
    except:
        print("Unexpected error:", sys.exc_info()[0])
        raise
        
    return db_info

#Module for fetching target protein data from demo db.

def tp_info(tp_id,base_url,tp_url):
    if tp_id !="":
        try:
            sess=requests.Session()
            url=base_url+tp_url+str(tp_id)
            response=sess.get(url,verify=False)
            tp_data=json.loads(response.text)
            sess.close()
        except:
            print("Unexpected error:", sys.exc_info()[0])
            raise
    else:
        pass
    
    return tp_data

#Module for fecthing actual protein data from demo db.
def ap_info(ap_id,base_url,ap_url,username,password):
    if ap_id !="":
        try:
            sess=requests.Session()
            url=base_url+ap_url+str(ap_id)
            response=sess.get(url,auth=HTTPBasicAuth(username,password),verify=False)
            apdata=response.text
            
            ap_tmp_file=open("ap_response.txt",'w',encoding='utf-8')
            ap_tmp_file.write(apdata)
            ap_tmp_file.close()
            
            ap_df=pd.read_csv("ap_response.txt",sep="\t")
            ap_df=ap_df.replace(np.nan, '', regex=True)
            sess.close()
            format_data=str(ap_df.at[0,'Format'])
        except:
            print("Unexpected error:", sys.exc_info()[0])
            raise
    else:
        pass
    
    return format_data

# Reading sample id from database.
def read_barcode(file):
    excel_file=op.load_workbook(file)
    ws=excel_file['Data Summary']
    rows=ws.max_row
    df=pd.read_excel(file)
    df.fillna(value="*",inplace=True)
    data={}
    
    count=1
    for idx in df.index:
        if str(df["Tube ID"][idx]).strip() != "*":
            data['labid']=str(df["Tube ID"][idx])
            db_row=oracleconn(data)
            
            val=db_row[5].split('.')
            
            
            count=count+1
            ws.cell(row=count,column=11,value=val[1])      #Batch
            ws.cell(row=count,column=12,value=val[0])      #PB
            
            
            creation_date=str(db_row[16]).split(" ")
            req_date = datetime.strptime(creation_date[0], "%Y-%m-%d").strftime('%m.%d.%Y')
            print (req_date)
            ws.cell(row=count,column=6,value=req_date)   #Request date
            ws.cell(row=count,column=27,value=db_row[5])   #Batch Id
            ws.cell(row=count,column=28,value=db_row[6])   #Batch Ref
            ws.cell(row=count,column=30,value=db_row[7])   #Amount
            ws.cell(row=count,column=31,value=db_row[8])   #Amount Units
            ws.cell(row=count,column=32,value=db_row[9])   #Concentration
            ws.cell(row=count,column=33,value=db_row[10])  #Concentration Units
            ws.cell(row=count,column=34,value=db_row[11])  #Solvent
        else:
            pass
     
    excel_file.save(file)
    
    return file

#Module for fetching data from Demo db.
def get_demodb_data(file,base_url,username,password):
    
    batch_url=r"Biologic/rest/proteinbatch?alias="
    ap_url=r"Biologic/rest/ap/"
    tp_url=r"demodbLookupService/tpProperties/"
    
    excel_file=op.load_workbook(file)
    ws=excel_file['Data Summary']
    rows=ws.max_row
    
    count=1
    for line  in range(2,rows+1):
        batch=str(ws.cell(row=line,column=11).value)
        if batch != "None":
            try:
                count=count+1
                sess=requests.Session()
                url=base_url+batch_url+str(batch)
                response=sess.get(url,auth=HTTPBasicAuth(username,password),verify=False)
                urldata=response.text

                tmp_file=open("response.txt",'w',encoding='utf-8')
                tmp_file.write(urldata)
                tmp_file.close()
                
                resp_df=pd.read_csv("response.txt",sep="\t")
                resp_df=resp_df.replace(np.nan, '', regex=True)
                ws.cell(row=count,column=9,value=str(resp_df.at[0,'Production dset ID'])) #col1
                ws.cell(row=count,column=10,value=str(resp_df.at[0,'Dept'])) #col2
                ws.cell(row=count,column=13,value=str(resp_df.at[0,'Target Protein ID'])) #col3 
                ws.cell(row=count,column=14,value=str(resp_df.at[0,'Actual Protein ID'])) #col4 
                ws.cell(row=count,column=15,value=str(resp_df.at[0,'Internal ID'])) #col5 
                ws.cell(row=count,column=16,value=str(resp_df.at[0,'Target Protein Name'])) #col6 
                ws.cell(row=count,column=22,value=str(resp_df.at[0,'Final Con [mg/mL]'])) #col7 
                ws.cell(row=count,column=23,value=str(resp_df.at[0,'Final Formulation'])) #col8 
                ws.cell(row=count,column=24,value=str(resp_df.at[0,'Storage Condition'])) #col9 
                ws.cell(row=count,column=25,value=str(resp_df.at[0,'Sample Yield [mg/L]'])) #col10
                ws.cell(row=count,column=26,value=str(resp_df.at[0,'RCode'])) #col11
                
                tp_data=tp_info(resp_df.at[0,'Target Protein ID'],base_url,tp_url)
                ws.cell(row=count,column=18,value=tp_data.get("A280_oxi",""))
                ws.cell(row=count,column=19,value=tp_data.get("isoelectric_point_oxi",""))
                ws.cell(row=count,column=21,value=tp_data.get("MW_oxidized",""))
                
                ap_format=ap_info(resp_df.at[0,'Actual Protein ID'],base_url,ap_url,username,password)
                if ap_format != "" or ap_format != "None" or ap_format != None:
                    ws.cell(row=count,column=20,value=ap_format)
                else:
                    ws.cell(row=count,column=20,value="")
                    
                sess.close()
                os.remove("response.txt")
                os.remove("ap_response.txt")
            except:
                print("Unexpected error:", sys.exc_info()[0])
                raise    
        else:
            pass
    
    excel_file.save(file)
    
# The following modules will generate data for data arrangment view.

def oraclebatch_ref1(batch_ref1,ws,rows):
    col1=[4,5,6,7,8,9,10,11,12,13,14,15] #cols
    row1=[5,6,7,8,9,10,11,12] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
             if count < rows-1:
                ws.cell(row=j,column=i,value=batch_ref1[count])
                count=count+1
    print ("BATCH_REF1 LENGTH",len(batch_ref1))
    print ("ROWS",rows)
    print ("COUNT",count)
    
def oraclebatch_ref2(batch_ref2,ws,rows):
    col1=[4,5,6,7,8,9,10,11,12,13,14,15] #cols
    row1=[20,21,22,23,24,25,26,27] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
             if count < rows-1:
                ws.cell(row=j,column=i,value=batch_ref2[count])
                count=count+1   
    
def oraclebatchid(batchid,ws,rows):
    col2=[21,22,23,24,25,26,27,28,29,30,31,32] #cols
    row2=[5,6,7,8,9,10,11,12] #rows
    
    count=0
    
    for i in col2:
        for j in row2:
            if count < rows-1:
                ws.cell(row=j,column=i,value=batchid[count])
                count=count+1
    
def oracletubeid(tubeid,ws,rows):
    col3=[38,39,40,41,42,43,44,45,46,47,48,49] #cols
    row3=[5,6,7,8,9,10,11,12] #rows
    
    count=0
    
    for i in col3:
        for j in row3:
            if count < rows-1:
                ws.cell(row=j,column=i,value=tubeid[count])
                count=count+1

def req_mw(mw,ws,rows):
    col1=[21,22,23,24,25,26,27,28,29,30,31,32] #cols
    row1=[20,21,22,23,24,25,26,27] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=mw[count])
                count=count+1
                
def scanner_free_text(scanner_data,ws,rows):
    col1=[38,39,40,41,42,43,44,45,46,47,48,49] #cols
    row1=[20,21,22,23,24,25,26,27] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=scanner_data[count])
                count=count+1                
                
                          
def oracleamount(oracleamt,ws,rows):
    col1=[4,5,6,7,8,9,10,11,12,13,14,15] #cols
    row1=[34,35,36,37,38,39,40,41] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=oracleamt[count])
                count=count+1
    
def request_pds(req_pds,ws,rows):
    col1=[21,22,23,24,25,26,27,28,29,30,31,32] #cols
    row1=[34,35,36,37,38,39,40,41] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=req_pds[count])
                count=count+1
            
def mosaicdb_conc(oracleconc,ws,rows):
    col1=[4,5,6,7,8,9,10,11,12,13,14,15] #cols
    row1=[49,50,51,52,53,54,55,56] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=oracleconc[count])
                count=count+1
            
def request_conc(req_conc,ws,rows):
    col1=[21,22,23,24,25,26,27,28,29,30,31,32] #cols
    row1=[49,50,51,52,53,54,55,56] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=req_conc[count])
                count=count+1
            
def mosaicdb_solvent(oraclesolvent,ws,rows):
    col1=[4,5,6,7,8,9,10,11,12,13,14,15] #cols
    row1=[64,65,66,67,68,69,70,71] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=oraclesolvent[count])
                count=count+1


def request_formulation(req_formulation,ws,rows):
    col1=[21,22,23,24,25,26,27,28,29,30,31,32] #cols
    row1=[64,65,66,67,68,69,70,71] #rows
    
    count=0
    
    for i in col1:
        for j in row1:
            if count < rows-1:
                ws.cell(row=j,column=i,value=req_formulation[count])
                count=count+1

#Creating data rack dataframe.

def data_rack(file):
    excel_file=op.load_workbook(file)
    df=pd.read_excel(file,sheet_name="Data Summary",header=0)
    df=df.replace(np.nan, '', regex=True)
    ws1=excel_file['Summary']
    rows=ws1.max_row
    ws=excel_file['Data View']
    
    batch_ref1=[]
    batch_ref2=[]
    batchid=[]
    tubeid=[]
    mw=[]
    oracleamt=[]
    req_pds=[]
    oracleconc=[]
    req_conc=[]
    oraclesolvent=[]
    req_formulation=[]
    scanner_ftext=[]
    
    for idx in df.index:
        batch_data=str(df['Batch Id'][idx])

        if batch_data != "" and batch_data is not None:
            try:
                batch_ref=batch_data.split('.')
                batch_ref1.apend(batch_ref[0])
                batch_ref2.apend(batch_ref[1])
            except:
                print("Unexpected error:", sys.exc_info()[0])
                raise    
        else:
            batch_ref1.apend("")
            batch_ref2.apend("")
            
        batch_id=str(df['Batch Id'][idx])
        tube_id=str(df['Tube ID'][idx])
        mw_kda=str(df['MW [kDa]'][idx])
        amt_mosaic=str(df['Amount'][idx])
        pds_req=str(df['PDS'][idx])
        conc_mosaic=str(df['Concentration'][idx])
        conc_req=str(df['Concentration (mg/ml)'][idx])
        sol_mosaic=str(df['Solvent'][idx])
        formul_req=str(df['Formulation'][idx])
        scanner_text=str(df['Free Text'][idx])
        
        batchid.apend(batch_id)
        tubeid.apend(tube_id)
        mw.apend(mw_kda)
        oracleamt.apend(amt_mosaic)
        req_pds.apend(pds_req)
        oracleconc.apend(conc_mosaic)
        req_conc.apend(conc_req)
        oraclesolvent.apend(sol_mosaic)
        req_formulation.apend(formul_req)
        scanner_ftext.apend(scanner_text)

    oraclebatch_ref1(batch_ref1,ws,rows)
    oraclebatch_ref2(batch_ref2,ws,rows)
    oraclebatchid(batchid,ws,rows)
    oracletubeid(tubeid,ws,rows)
    req_mw(mw,ws,rows)
    scanner_free_text(scanner_ftext,ws,rows)
    oracleamount(oracleamt,ws,rows)
    request_pds(req_pds,ws,rows)
    mosaicdb_conc(oracleconc,ws,rows)
    request_conc(req_conc,ws,rows)
    mosaicdb_solvent(oraclesolvent,ws,rows)
    request_formulation(req_formulation,ws,rows)
   
    excel_file.save(file) #Saving output file
    
######################################################## MAIN FUNCTION ##################################################  
if __name__ == "__main__":
    file=r"/local_system/dummy_path/xyz.xlsx" #Path of Input file
    base_url=r"https://www.dummyurl.com" #Base Url of database
    username="dummy_user_name" #User name of database
    password="dummy_password" #password of database
    output_file=read_barcode(file) #Invoking oracle connection module.
    get_demodb_data(output_file,base_url,username,password) #Invoking customised database for other metedata.
    data_rack(file) #Invoking creating rack view of above metadata.


# In[ ]:




