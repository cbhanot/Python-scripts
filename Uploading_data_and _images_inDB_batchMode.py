#!/usr/bin/env python
# coding: utf-8

# In[4]:


####### DESCRIPTION#########
#This code uploads data and modified image files in a customised database in batch mode by identifying newer files in a 
#given directory and return the output files.

"""
The code reads the data from input file, creates a tsv file on the fly and modifies the image files from the path
given and then submits the data and images in the database and returns the submission id along with submission status by 
parsing the response from database and writing the same in output file and emailing the results on user mail id.

The overall idea behind writing this code was to reduce manual intervention by using the power and rich libraries of
python in order to increase the data submission speed with increased efficiency along with status ofsubmission.
"""

#Input: Excel sheet containing data and path of images present in local system/server.
#Output: Returning the excel sheet populated with the above data along with submission entry id from the db and 
#upload status.


### Python libraries used for achieving objective#########

import sys
import xlrd as xl
from xlutils.copy import copy
import xlwt as wt
import regex as re
import operator
import time
from datetime import datetime
import os
import os.path
from os import path
import subprocess
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
import smtplib
import openpyxl
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import json
import requests
from requests.auth import HTTPBasicAuth
###########################################################################################################################
# Module to read user credentials and email info from configuration file for uploading and sending mail to user.

def config_file(file):
    with open(file) as f:
        credentials=json.load(f)
    f.close()
    return credentials
############################################################################################################################
#Module for sending emails of process completion t user's.

def send_mail(receiver_mail,subject,msg):
    mail= smtplib.SMTP('smtp.gmail.com',587)
    mail.ehlo()
    mail.starttls()
    email="sender@gmail.com"   #Dummy mail id 
    password="sender"          #Dummy password
    mail.login(email, password)
    content = 'Subject:{}\n\n{}'.format(subject,msg)
    mail.sendmail(email,receiver_mail,content)
    mail.close()
#############################################################################################################################
#Module for creating a tsv on fly for data sumbission in database.

def create_tsv(col_list,ID_idx,path_to_temp,row,lr,data2,img_loc,ID,Name,MT):
    now=datetime.now()
    timestamp=datetime.timestamp(now)
    
    h2=""
    h1="SUB ID"+ "\t"+"SUB Name"+"\t"+"Analyzed Entity Type" + "\t"+ "Analyzed Entity ID"+ "\t" + "Analyzed Entity Name"+ "\t"+ "Laboratory Result Type" + "\t"+ "Measurement Type"
    for k,v in col_list.items():
        if v > 1 and v < ID_idx:   
            h2=h2+"\t"+k
    tsv_header=h1+h2+"\t"+"Reference to Image-Original"
    
    TSVfilePath=str(path_to_temp)+ "demodbImageUploaderTMP"+ str(timestamp) + ".tsv"
    TSV= open(TSVfilePath,"w")
    TSV.write(tsv_header+"\n")
    
    data1=str(lr)+"\t\t"+"Protein Batch"+"\t"+str(ID)+"\t"+str(Name)+"\t\t"+str(MT)+"\t"
    tsv_data=data1+data2+"\t"+img_loc   
    TSV.write(tsv_data) 
    TSV.close()
    
    return TSVfilePath
############################################################################################################################    
#Module for uploading data present in tsv file in database using credentials from configuration file.

def upload_tsv(file_path,user_name,passwd):
    if os.path.isfile(file_path):
        with open(file_path, 'rb') as file:
            headers={'Content-Type':'text/tsv', 'charset':'ISO-8859-4'}
            response = requests.post('https://demodb-dev.xyz.com/Biologic/ws/rest/laboratoryres',auth=HTTPBasicAuth(user_name,passwd),headers=headers, data=file.read() ,verify=False)
############################################################################################################################
#Module for returning any error msg received from server upon request.
def error_response(val):
    error={'400':'ERROR:Malformed syntax','401':'ERROR: Unauthorized access caused due to user authentication','402':'ERROR:Payment Required-This code is reserved for future use','403':'ERROR:Forbidden','404':'ERROR:Not Found:The server has not found anything matching the Request-URI','405':'ERROR:Method Not Allowed','406':'ERROR:Not Acceptable','407':'ERROR:Proxy Authentication Required','408':'ERROR:Request Timeout','409':'ERROR:Conflict','410':'ERROR:The requested resource is no longer available at the server and no forwarding address is known'
           ,'411':'ERROR: The server refuses to accept the request without a defined Content-Length','412':'ERROR:The precondition given in one or more of the request-header fields evaluated to false when it was tested on the server','413':'ERROR:The server is refusing to process a request because the request entity is larger than the server is willing or able to process','414':'ERROR:The server is refusing to service the request because the Request-URI is longer than the server is willing to interpret'
           ,'415':'ERROR:The server is refusing to service the request because the entity of the request is in a format not supported by the requested resource for the requested method','416':'ERROR:Requested Range Not Satisfiable','417':'ERROR:The expectation given in an Expect request-header field could not be met by this server','500':'ERROR:Internal Server Error','501':'ERROR:The server does not support the functionality required to fulfill the request','502':'ERROR:Bad Gateway','503':'ERROR:Service Unavailable'
           ,'504':'ERROR:Gateway Timeout',505:'ERROR:HTTP Version Not Supported'}
    return error.get(val)
############################################################################################################################
# Module for parsing xml response received from database after submission.

def parse_xml(response):
    root=ET.fromstring(response)
    val=root.attrib['responseStatus']

    if val == "200":
        pattern=re.compile(r"Sub-\d+")
        cnt=0
        for branch in root:
            for leaf in branch:
                cnt=cnt+1
                if cnt == 1:
                    data=[]
                    for v in leaf.attrib.values():
                        data.append(v)
                else:
                    cnt=0
                for entity in leaf:
                    if pattern.search(entity.text):
                        data.append(entity.text)
        return data
    else:
        data=[]
        data.append(error_response(val))
        return data
#############################################################################################################################
# Module for modifing image files.

def texonImage(imgfilepath, PB_ID, PB_NAME):
#    Image format extension
    imgformat = ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif', '.rgb', '.pbm', '.pgm', '.ppm', '.rast', '.xbm', '.webp', '.exr']
#    Validating the image file
    if imgfilepath == "" or imgfilepath == None or os.path.exists(imgfilepath) == False or imgfilepath.lower().endswith((tuple(imgformat))) == False:
        print("Invalid")    
    else:      
#    Parsing the image from input file
        image = Image.open(imgfilepath).convert('RGB')
#    initialise the drawing context with
#    the image object as background
        draw = ImageDraw.Draw(image)
#    create font object with the font file and specify
#    desired size
        font = ImageFont.truetype('arial.ttf', size=15)
#    starting position of the message
        (x, y) = (45, 2)    
#    Validating PB_NAME exits
        if PB_NAME == None:
            message = PB_ID
        else:
            message = PB_ID + " | " + PB_NAME    
        color = 'rgb(0, 0, 0)' # black color
#    draw the message on the background
        draw.text((x, y), message, fill=color, font=font)    
#    Saving the image
        image.save(imgfilepath)
###########################################################################################################################
# Module for uploaing image file in database.

def upload_image(img_path,ID,Name,user_name,passwd):
    if os.path.isfile(image):
        
        if ID != "":
            
            with open(image, 'rb') as img:
                name_img= os.path.basename(image)
            
                img_url= f"https://demodb-dev.xyz.com/demodbImageUpload/upload"                 f"?folder={ID}"                 f"&filename={name_img}"                 f"&secret=XXXXX"
            
                response=requests.post(img_url, auth=HTTPBasicAuth(user_name,passwd), data=img.read(),verify=False)
        else:
            
            with open(image, 'rb') as img:
                name_img= os.path.basename(image)
            
                img_url= f"https://demodb-dev.xyz.com/demodbImageUpload/upload"                 f"?folder={Name}"                 f"&filename={name_img}"                 f"&secret=XXXXX"
            
                response=requests.post(img_url, auth=HTTPBasicAuth(user_name,passwd), data=img.read(),verify=False)
    else:
        print("File is not valid")
###########################################################################################################################

################################################ MAIN FUNCTION ############################################################

if __name__ == "__main__" :
    
    PATH=""   #Path to input directory
    tmp_folder="" #Path to temp folder for storing tmp tsv files
                                     
    config="" #Path to configuration file.
    
    DIR=os.scandir(PATH)

    for file in DIR:
        if file.is_file():
            xls=file.name
            if not xls.endswith("_INPROGRESS.xlsx") and not xls.endswith("_COMPLETED.xlsx"):  #Checking for new files in directory.
                temp1=xls.split(".")
                src=str(PATH)+"//"+xls
                dest1=str(PATH)+"//"+str(temp1[0])+"_INPROGRESS."+str(temp1[1])
                os.rename(src,dest1)    #Renaming original file with INPROGRESS file
    
                print ("Execution started")
        
                worksheet_name=dest1                                                      

                wb=xl.open_workbook(worksheet_name) #Reading excel file
                pxl=load_workbook(worksheet_name) #Reading excel for writing data in it.
    
                a=pxl.sheetnames #getting sheet names in list
    
                ws=pxl[a[0]] #creating sheet object for writing in first sheet of workbook

                sheet=wb.sheet_by_index(0) #Accessing sheet by index

                nrows=sheet.nrows #getting no of rows from sheet
                ncols=sheet.ncols # getting no of columns from sheet

                col_list={} #creating a empty dictionary for storing columns and their index

                if nrows >3:
    
                    MT=sheet.cell_value(0,2) #storing the measurement type
                    head=[] #creating an empty list for storing column name 
                    for colname in range(ncols): #Loop over columns
                        columns=re.sub('\[.*\]','',sheet.cell_value(2,colname)) #subsituting [] pattern with nothing
                        col_list[columns]=colname #storing values in dictionary
                        head.append(columns) #appending columns name in head list
                    head="\t".join(head) # joining list items as string
    
                    idx_id=col_list['PB_ID']        #getting index location for PB_ID
                    idx_name=col_list['PB Name']    #getting index location for PB Name
                    idx_il=col_list['Image Location']   #getting index location for Image Location
    
                    if 'PB_ID' in col_list and 'PB Name' in col_list and 'Image Location' in col_list: #Checking the presence of mentioned columns
        
                        for r in range(3,nrows):   #Iterating over rows from 3 till last row in file
            
                            pattern=re.compile(r'(PB-\d+)') #defining pattern for PB_ID
                            ID=sheet.cell_value(r,col_list.get('PB_ID')) #getting the value of PB_ID in a row and storing in ID
                            Name=sheet.cell_value(r,col_list.get('PB Name'))#getting the values of PB Name in a row and storing in Name
            
                            if pattern.search(ID) or (ID == "" and Name != ""): #condition to check whether PB_ID is in correct format or if PB_ID is null than PB Name should be present.
            
                                img_loc=sheet.cell_value(r,col_list.get('Image Location')) # storing image location from the file for each row
                                test=os.path.isfile(img_loc)#checking the validity of image file.
                
                                if img_loc != "" and  test == True: #condition to check if image location is not blank and the image path is valid.
                                    SUB_ID=["" if sheet.cell_value(r,col_list.get('SUB-ID')) == "" else sheet.cell_value(r,col_list.get('SUB-ID'))]
                                    SUB_ID="".join(SUB_ID)    #Capturing SUB_ID if already present in input file
                        
                                    texonImage(img_loc,ID,Name)  #Adding text on Image function
                        
                                    data2=[]
                                    for c in range(2,idx_id):
                                        val=str(sheet.cell_value(r,c))    #Generating string of data present in a particular row of sheet.
                                        val=val.replace(',','.')
                                        data2.append(val)
                                    data2="\t".join(map(str,data2))
                    
                                    tsv=create_tsv(col_list,idx_id,tmp_folder,r,SUB_ID,data2,img_loc,ID,Name,MT)  #Invoking create_tsv module for creating TSV file
                        
                                    xml_response=upload_tsv(tsv,cred['u_name'],cred['password']) #Invoking upload_tsv module for uploading data in demodb.
                        
                                    response=parse_xml(xml_response) #Invoking parse_xml module for processing response
                        
                                    img_upload_response=upload_image(img_loc,ID,Name,cred['u_name'],cred['password'])# Invoking upload_image module for uploading dimage in demodb
                        
                                    ############################################################################
                                    #          Filling values of reponses in sheet
                                    ws.cell(row=(r+1),coulmn=1,value=response[1])   # Filling SUB ID in sheet
                                    ws.cell(row=(r+1),column=(idx_il+2),value=image_upload_response)
                                    ws.cell(row=(r+1),column=(idx_il+3),value=response[0])
                                    ws.cell(row=(r+1),column=(idx_il+4),value=response[0])
                                    ###########################################################################
                        
                                elif img_loc == "" or test == False:
                    
                                    SUB_ID=["" if sheet.cell_value(r,col_list.get('SUB-ID')) == "" else sheet.cell_value(r,col_list.get('SUB-ID'))]
                                    SUB_ID="".join(SUB_ID)    #Capturing SUB_ID if already present in input file
                    
                                    data2=[]
                                    for c in range(2,idx_id):          #Generating string of data present in a particular row of sheet.
                                        val=str(sheet.cell_value(r,c)).replace(",",".")
                                        data2.append(val)
                                    data2="\t".join(map(str,data2))
                    
                                    tsv=create_tsv(col_list,idx_id,tmp_folder,r,SUB_ID,data2,img_loc,ID,Name,MT) #Invoking create_tsv module for creating TSV file
                        
                                    xml_response=upload_tsv(tsv,cred['u_name'],cred['password']) #Invoking upload_tsv module for uploading data in demodb.
                                    
                                    response=parse_xml(xml_response) #Invoking parse_xml module for processing response 
                        
                                    ############################################################################
                                    #           Filling values of reponses in sheet
                                    ws.cell(row=(r+1),coulmn=1,value=response[1])   # Filling SUB ID in sheet
                                    ws.cell(row=(r+1),column=(idx_il+2),value="ERROR")
                                    ws.cell(row=(r+1),column=(idx_il+3),value=response[0])
                                    ws.cell(row=(r+1),column=(idx_il+4),value=response[0])
                                    ############################################################################
                                else:
                                    ws.cell(row=(r+1),column=(idx_il+2),value="ERROR")
                                    ws.cell(row=(r+1),column=(idx_il+3),value="")
                                    ws.cell(row=(r+1),column=(idx_il+4),value="")
                        
                            elif ID == "" and Name == "":
                                ws.cell(row=(r+1),column=(idx_il+2),value="")
                                ws.cell(row=(r+1),column=(idx_il+3),value="")
                                ws.cell(row=(r+1),column=(idx_il+4),value="PB_ID and/or PB Name must be specified!")
                    
                            else:
                                ws.cell(row=(r+1),column=(idx_il+2),value="ERROR")
                                ws.cell(row=(r+1),column=(idx_il+3),value="")
                                ws.cell(row=(r+1),column=(idx_il+4),value="")
                    
                    pxl.save(worksheet_name) #Saving worksheet after entering data 
                    final_dest=str(PATH)+"//"+str(temp1[0])+"_COMPLETED."+str(temp1[1]) # Creating file for complete process
                    os.rename(dest1,final_dest)  #Renaming INPROGRESS file to COMPLETED
                                                                      
                else:
                    sub="FILE TEMPLATE ERROR"
                    msg="There is some issue with FILE TEMPLATE, Kindly check headers in file"
                    send_mail(cred['email'],sub,msg)
                
            else:
                sub="DATA ERROR"
                msg="Do not process empty file,kindly enter data."
                send_mail(cred['email'],sub,msg)
        print("Execution completed")


# In[ ]:




